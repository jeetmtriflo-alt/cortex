from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
import io
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_file_by_format(file, filename):
    """Read file based on its format (CSV or XLSX)"""
    file_ext = filename.rsplit('.', 1)[1].lower()
    
    if file_ext == 'csv':
        return pd.read_csv(file)
    elif file_ext in ['xlsx', 'xls']:
        return pd.read_excel(file)
    else:
        raise ValueError(f"Unsupported file format: {file_ext}")


def get_column_name(df, column_str):
    """Get column name from dataframe by index, Excel letter, or column name"""
    # Try column name first
    if column_str in df.columns:
        return column_str
    
    # Try numeric index
    try:
        col_idx = int(column_str)
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]
    except ValueError:
        pass
    
    # Try Excel column letter (A=0, B=1, etc.)
    try:
        column_str_upper = column_str.upper()
        result = 0
        for char in column_str_upper:
            result = result * 26 + (ord(char) - ord('A') + 1)
        col_idx = result - 1
        if 0 <= col_idx < len(df.columns):
            return df.columns[col_idx]
    except:
        pass
    
    raise ValueError(f"Column '{column_str}' not found")


def filter_printer_hostnames(hostnames):
    """Filter out hostnames containing HP, canon, NPI, or printer (case-insensitive)"""
    if not hostnames:
        return []
    
    # Keywords to filter out (case-insensitive)
    filter_keywords = ['hp', 'canon', 'npi', 'printer']
    
    filtered = []
    for hostname in hostnames:
        hostname_lower = str(hostname).lower()
        # Check if hostname contains any of the filter keywords
        if not any(keyword in hostname_lower for keyword in filter_keywords):
            filtered.append(hostname)
    
    return filtered


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/compare', methods=['POST'])
def compare_files():
    try:
        # Check if files are present
        if 'source_file' not in request.files or 'reference_file' not in request.files:
            return jsonify({'error': 'Please upload both source and reference files'}), 400
        
        source_file = request.files['source_file']
        reference_file = request.files['reference_file']
        endpoints_file = request.files.get('endpoints_file')  # Optional third file
        
        # Get column names/indices (default to 'Hostname' for source and 'Endpoint Name' for reference)
        source_col_str = request.form.get('source_column', 'Hostname')
        reference_col_str = request.form.get('reference_column', 'Endpoint Name')
        
        # Validate files
        if source_file.filename == '' or reference_file.filename == '':
            return jsonify({'error': 'Please select both files'}), 400
        
        if not (allowed_file(source_file.filename) and allowed_file(reference_file.filename)):
            return jsonify({'error': 'Invalid file types. Please upload CSV or XLSX files'}), 400
        
        # Read files based on their format
        source_df = read_file_by_format(source_file, source_file.filename)
        reference_df = read_file_by_format(reference_file, reference_file.filename)
        
        # Fill NaN
        source_df = source_df.fillna('')
        reference_df = reference_df.fillna('')
        
        # Get column names
        try:
            source_col_name = get_column_name(source_df, source_col_str)
        except ValueError as e:
            available_cols = ', '.join([f"'{col}'" for col in source_df.columns])
            return jsonify({'error': f'Source file column error: {str(e)}. Available columns: {available_cols}'}), 400
        
        try:
            reference_col_name = get_column_name(reference_df, reference_col_str)
        except ValueError as e:
            available_cols = ', '.join([f"'{col}'" for col in reference_df.columns])
            return jsonify({'error': f'Reference file column error: {str(e)}. Available columns: {available_cols}'}), 400
        
        # Normalize columns (matching reference script exactly)
        reference_endpoints = (
            reference_df[reference_col_name]
            .astype(str)
            .str.strip()
            .str.lower()
        )
        
        source_hostnames = (
            source_df[source_col_name]
            .astype(str)
            .str.strip()
            .str.lower()
        )
        
        # Remove empty hostnames first
        source_df_clean = source_df[source_hostnames != '']
        
        # Recompute cleaned hostnames
        source_hostnames_clean = (
            source_df_clean[source_col_name]
            .astype(str)
            .str.strip()
            .str.lower()
        )
        
        source_total = len(source_hostnames_clean)
        reference_total = len(reference_endpoints[reference_endpoints != ''].unique())
        
        # Find hostnames NOT present in reference file (case-insensitive)
        not_matching = source_df_clean.loc[
            ~source_hostnames_clean.isin(reference_endpoints),
            source_col_name
        ]
        
        unique_in_source = not_matching.tolist()
        
        # Filter out printer-related hostnames (HP, canon, NPI, printer)
        filter_printers_val = request.form.get('filter_printers', 'true')
        # Handle checkbox values: 'on' (checked), 'true' (explicit), or 'false'/'off' (unchecked)
        filter_printers = filter_printers_val.lower() in ['true', 'on', '1', 'yes']
        if filter_printers:
            unique_in_source_filtered = filter_printer_hostnames(unique_in_source)
            filtered_count = len(unique_in_source) - len(unique_in_source_filtered)
        else:
            unique_in_source_filtered = unique_in_source
            filtered_count = 0
        
        # Also find endpoints in reference that are NOT in source (for Column F)
        source_hostnames_set = set(source_hostnames_clean)
        reference_endpoints_clean = reference_endpoints[reference_endpoints != '']
        unique_in_reference = [
            ref_val for ref_val in reference_df[reference_col_name].astype(str).str.strip() 
            if ref_val.lower() not in source_hostnames_set and ref_val != ''
        ]
        
        # Process endpoints file if provided (for Column F - Endpoints without Cortex Agent)
        endpoints_without_agent_count = len(unique_in_reference)
        if endpoints_file and endpoints_file.filename != '':
            try:
                endpoints_df = read_file_by_format(endpoints_file, endpoints_file.filename)
                endpoints_df = endpoints_df.fillna('')
                # Try to get the column - default to first column
                endpoints_col = get_column_name(endpoints_df, request.form.get('endpoints_column', '0'))
                endpoints_list = endpoints_df[endpoints_col].astype(str).str.strip()
                endpoints_list = endpoints_list[endpoints_list != ''].tolist()
                endpoints_without_agent_count = len(endpoints_list)
            except Exception as e:
                # If processing fails, use the calculated value
                pass
        
        return jsonify({
            'success': True,
            'source_total': int(source_total),
            'reference_total': int(reference_total),
            'unique_count': len(unique_in_source_filtered),
            'unique_hostnames': unique_in_source_filtered,
            'filtered_count': filtered_count,
            'original_unique_count': len(unique_in_source),
            'unique_in_reference_count': endpoints_without_agent_count
        })
    
    except Exception as e:
        import traceback
        return jsonify({'error': f'Error during comparison: {str(e)}\n{traceback.format_exc()}'}), 500


@app.route('/download', methods=['POST'])
def download_results():
    try:
        data = request.json
        unique_hostnames = data.get('hostnames', [])
        export_format = data.get('format', 'csv')  # 'csv' or 'xlsx'
        
        # Get statistics for Excel template format
        source_total = data.get('source_total', 0)
        reference_total = data.get('reference_total', 0)
        unique_count = data.get('unique_count', 0)
        unique_in_reference_count = data.get('unique_in_reference_count', 0)
        
        output = io.BytesIO()
        
        if export_format.lower() == 'xlsx':
            # Create Excel file in template format
            wb = Workbook()
            ws = wb.active
            ws.title = "Cortex Coverage Report"
            
            # Header row with styling
            headers = [
                'Date',
                'Total Endpoints identified by script',
                'Total Endpoints registered with Cortex',
                'Endpoints without Cortex Agent (Script)',
                'Total endpoints identified by Corte',
                'Endpoints without Cortex Agent (Cortex)'
            ]
            
            # Style for header row
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Get today's date in format like "1-Jan-26"
            today = datetime.now()
            # Format: day without leading zero, abbreviated month, 2-digit year
            day = today.day
            month_abbr = today.strftime('%b')
            year_2digit = today.strftime('%y')
            date_str = f'{day}-{month_abbr}-{year_2digit}'  # e.g., "1-Jan-26"
            
            # Data row
            row_data = [
                date_str,
                source_total,
                reference_total,
                unique_count,
                reference_total,  # Same as Column C
                unique_in_reference_count
            ]
            
            # Write data row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                if col_idx == 1:  # Date column
                    cell.alignment = Alignment(horizontal='left')
                else:  # Number columns
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            column_widths = [12, 35, 35, 35, 35, 35]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
            
            # Create second sheet for Unique Hostnames
            ws_hostnames = wb.create_sheet("Unique Hostnames")
            
            # Headers for unique hostnames sheet
            ws_hostnames.cell(row=1, column=1, value="Date").font = header_font
            ws_hostnames.cell(row=1, column=1).fill = header_fill
            ws_hostnames.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            ws_hostnames.cell(row=1, column=2, value="Hostname").font = header_font
            ws_hostnames.cell(row=1, column=2).fill = header_fill
            ws_hostnames.cell(row=1, column=2).alignment = Alignment(horizontal='center')
            
            # Get desktop path and unique hostnames file
            desktop_path = Path.home() / "Desktop"
            unique_hostnames_file = desktop_path / "unique_hostnames.xlsx"
            
            # Read existing unique hostnames if file exists
            existing_data = []
            if unique_hostnames_file.exists():
                try:
                    existing_wb = load_workbook(unique_hostnames_file)
                    if "Unique Hostnames" in existing_wb.sheetnames:
                        existing_ws = existing_wb["Unique Hostnames"]
                        for row in existing_ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and row[1]:  # Date and Hostname
                                existing_data.append((row[0], row[1]))
                except Exception as e:
                    # If file is corrupted or can't be read, start fresh
                    pass
            
            # Add existing data to new sheet
            row_num = 2
            for date_val, hostname_val in existing_data:
                ws_hostnames.cell(row=row_num, column=1, value=date_val)
                ws_hostnames.cell(row=row_num, column=2, value=hostname_val)
                row_num += 1
            
            # Add today's unique hostnames
            today_date_str = today.strftime('%Y-%m-%d')
            for hostname in unique_hostnames:
                ws_hostnames.cell(row=row_num, column=1, value=today_date_str)
                ws_hostnames.cell(row=row_num, column=2, value=hostname)
                row_num += 1
            
            # Set column widths for hostnames sheet
            ws_hostnames.column_dimensions['A'].width = 15
            ws_hostnames.column_dimensions['B'].width = 50
            
            # Save unique hostnames to desktop file
            try:
                wb_hostnames = Workbook()
                ws_hostnames_desktop = wb_hostnames.active
                ws_hostnames_desktop.title = "Unique Hostnames"
                
                # Headers
                ws_hostnames_desktop.cell(row=1, column=1, value="Date").font = header_font
                ws_hostnames_desktop.cell(row=1, column=1).fill = header_fill
                ws_hostnames_desktop.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                ws_hostnames_desktop.cell(row=1, column=2, value="Hostname").font = header_font
                ws_hostnames_desktop.cell(row=1, column=2).fill = header_fill
                ws_hostnames_desktop.cell(row=1, column=2).alignment = Alignment(horizontal='center')
                
                # Add all data (existing + new)
                row_num_desktop = 2
                for date_val, hostname_val in existing_data:
                    ws_hostnames_desktop.cell(row=row_num_desktop, column=1, value=date_val)
                    ws_hostnames_desktop.cell(row=row_num_desktop, column=2, value=hostname_val)
                    row_num_desktop += 1
                
                for hostname in unique_hostnames:
                    ws_hostnames_desktop.cell(row=row_num_desktop, column=1, value=today_date_str)
                    ws_hostnames_desktop.cell(row=row_num_desktop, column=2, value=hostname)
                    row_num_desktop += 1
                
                ws_hostnames_desktop.column_dimensions['A'].width = 15
                ws_hostnames_desktop.column_dimensions['B'].width = 50
                
                wb_hostnames.save(unique_hostnames_file)
            except Exception as e:
                # If saving to desktop fails, continue with export
                pass
            
            wb.save(output)
            output.seek(0)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'Cortex_Coverage_Report_{today.strftime("%Y%m%d")}.xlsx'
        else:
            # Export to CSV format (simple list)
            if not unique_hostnames:
                return jsonify({'error': 'No hostnames to export'}), 400
            
            df = pd.DataFrame(unique_hostnames, columns=['Unique_Hostnames'])
            df.to_csv(output, index=False)
            output.seek(0)
            mimetype = 'text/csv'
            filename = 'unique_hostnames.csv'
        
        return send_file(
            output,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        import traceback
        return jsonify({'error': f'Error exporting results: {str(e)}\n{traceback.format_exc()}'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

